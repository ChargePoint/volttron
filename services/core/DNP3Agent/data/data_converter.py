# -*- coding: utf-8 -*- {{{
# vim: set fenc=utf-8 ft=python sw=4 ts=4 sts=4 et:
#
# Copyright 2018, 8minutenergy / Kisensum.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
#
# Neither 8minutenergy nor Kisensum, nor any of their
# employees, nor any jurisdiction or organization that has cooperated in the
# development of these materials, makes any warranty, express or
# implied, or assumes any legal liability or responsibility for the accuracy,
# completeness, or usefulness or any information, apparatus, product,
# software, or process disclosed, or represents that its use would not infringe
# privately owned rights. Reference herein to any specific commercial product,
# process, or service by trade name, trademark, manufacturer, or otherwise
# does not necessarily constitute or imply its endorsement, recommendation, or
# favoring by 8minutenergy or Kisensum.
# }}}

import csv
import json
# This library must be loaded ("pip install openpyxl") before executing this script.
from openpyxl import load_workbook
import os

from data_converter_constants import *


class Converter(object):
    """
        Transform DNP3 point and function definition data from spreadsheet into config files.

        The source of the point definitions is the MESA-ESS "DNP3 Point Spreadsheet"
        at http://mesastandards.org/mesa-ess-2016/. An .xls version of that data
        is checked into this project as mesa_points.xls.

        This script builds a point-definitions file, mesa_points.config,
        based on the contents of four of the spreadsheet's worksheets:

            BI (Binary Input point definitions)
            BO (Binary Output point definitions)
            AI (Analog Input point definitions)
            AO (Analog Output point definitions)

        The resulting mesa_points.config files defines various properties for each point definition.
        See points.py for descriptions of point definition properties.
    """

    def __init__(self):
        self.point_defs = []            # Flat list of all point definitions
        self.points_by_name = {}        # Dictionary by name of all point definitions
        self.function_defs = {}         # Dictionary of function definitions, keyed by function name
        # Next available index for new point creation, by data type
        self.next_index = {
            'AI': None,
            'AO': None,
            'BI': None,
            'BO': None,
        }

    def get_next_index(self, data_type_name):
        """Return the next available point index for the data type. Increment the next index value."""
        next_index = self.next_index[data_type_name]
        if next_index is None:
            # Lazily initialize next_index. Find the highest index in use so far.
            next_index = 0
            for pt in self.point_defs:
                if pt['data_type'] == data_type_name:
                    this_index = pt['index']
                    if this_index > next_index:
                        next_index = this_index
                        if 'type' in pt and pt['type'] == 'array':
                            # Advance an extra amount if the highest point is an array
                            next_index += len(pt['array_points']) * MAX_ARRAY_POINTS - 1
            next_index += 1                                      # This is the index to assign
        self.next_index[data_type_name] = next_index + 1         # This is the next available index
        return next_index

    @staticmethod
    def scrubbed_cell_value(cell_val):
        if cell_val is None:
            return ''
        elif type(cell_val) == long or type(cell_val) == float:
            return str(cell_val)
        else:
            cell_val = cell_val.encode('ascii', errors='ignore').decode()
            # The following line should remove newlines and carriage returns, but it didn't seem to do so.
            # No worries -- they get removed later anyway.
            cell_val = cell_val.replace('\n', ' ').replace('\r', ' ')
            return cell_val

    def create_point_csvs_from_excel(self):
        """Export data sheets from xlsx to csv files."""
        workbook = load_workbook(POINT_INPUT_FILE, data_only=True)
        # NOTE: To get this working properly, I had to hand-edit the .xlsx file, removing the
        # newline between 'Point' and 'Index' in the column header of each worksheet.
        # Otherwise they're read as two separate lines, and the all-important column header
        # row isn't read properly when the workbook is loaded.
        for worksheet_name in DATA_TYPES.keys():
            with open(''.join([worksheet_name, '.csv']), 'wb') as output_file:
                wr = csv.writer(output_file, quoting=csv.QUOTE_ALL)
                sheet = workbook[worksheet_name]
                for idx, row in enumerate(sheet.iter_rows()):
                    try:
                        wr.writerow([self.scrubbed_cell_value(cell.value) for cell in row])
                    except Exception as err:
                        print('Omitting row {} of worksheet {}: {}'.format(idx, worksheet_name, err))

    def add_point(self, point_def):
        self.point_defs.append(point_def)
        if point_def['name'] in self.points_by_name:
            raise ValueError('Duplicate point {}'.format(point_def))
        self.points_by_name[point_def['name']] = point_def

    def load_point(self, data_type_name, row_num, row, default_row_values):
        """Transform data from the input CSV row into a more consumable set of JSON data."""
        try:
            point_def = extra_point_data(data_type_name, row_num)
            if point_def.get('operation', None) == 'skip':
                return None
            if 'index' not in point_def:
                # Remove the Point Index's two-character data-type prefix. Confirm that the result is an integer.
                index = row.get('Point Index', '').replace('\n', ' ').replace('\r', ' ')
                index_name = index[0:2]
                if index_name != data_type_name:
                    raise ValueError('Unexpected data type name in point index = ', index_name)
                point_def['index'] = int(index[2:])
            if 'name' not in point_def:
                # Remove embedded newlines in the name.
                name = row.get('Unique String', '').replace('\n', ' ').replace('\r', ' ')
                if len(name) < 1 or name in BAD_POINT_NAMES:
                    # The row has a bad or nonexistent Unique String. Invent one.
                    name = data_type_name + '.' + row['Point Index']
                point_def['name'] = name
            point_def['data_type'] = data_type_name
            point_def['group'] = default_row_values['group']
            point_def['variation'] = default_row_values['variation']
            if row.get('Name / Description', None):
                point_def['description'] = row['Name / Description']
            if row.get('Units', None):
                point_def['units'] = row['Units']
            self.add_point(point_def)
        except ValueError as exc:
            # Skip the row
            err_msg = '\tBad {} row data: row_num={}, point_index={}, unique_string={}, exc={}'
            print(err_msg.format(data_type_name,
                                 row_num,
                                 row.get('Point Index', ''),
                                 row.get('Unique String', ''),
                                 exc))

    def create_extra_output_point(self, data_type_name, point_name):
        next_index = self.get_next_index(data_type_name)
        new_point = {
            'data_type': data_type_name,
            'index': next_index,
            'name': point_name,
            'group': DATA_TYPES[data_type_name]['group'],
            'variation': DATA_TYPES[data_type_name]['variation']
        }
        self.add_point(new_point)
        return new_point

    def get_output_point_by_name(self, point_name):
        """Look up an output point by point_name, returning None if not found."""
        pt = self.points_by_name.get(point_name, None)
        if pt is None:
            print('No point found with name {}'.format(point_name))
            return None
        else:
            if pt['group'] == 40 or pt['group'] == 10:
                return pt           # The point is an output point. Return it.
            else:
                # The point is not an output point. Create a new output point and return it.
                data_type_name = 'BO' if pt['group'] == 1 else 'AO'
                return self.create_extra_output_point(data_type_name, point_name + '.out')

    def create_function_step(self, func, row):
        """Create a function step."""
        function_name = func['name']
        step_number = int(row['Step '])
        # Substitute a different name for the step if necessary.
        point_name = extra_data_for_function_step(function_name, step_number, 'name')
        if point_name is None:
            point_name = row['IEC 61850 '].rstrip()
        # Confirm that the step's name corresponds to a point definition name.
        point_def = self.get_output_point_by_name(point_name)
        if point_def is not None:
            step = {
                'step_number': step_number,
                'point_name': point_def['name'],
                'description': row['Description ']
            }
            if 'M/O/C ' in row:
                optional_flag = row['M/O/C '].rstrip()
                if optional_flag:
                    if optional_flag in ['M', 'O', 'C']:
                        step['optional'] = optional_flag[:1]
                    else:
                        raise ValueError('Bad optional flag = {}'.format(optional_flag))
            if 'Function Codes ' in row:
                fcodes = row['Function Codes ']
                if fcodes:
                    fcodes_to_save = [FCODE_MAP[fc] for fc in fcodes.split(',') if FCODE_MAP[fc]]
                    if len(fcodes_to_save) > 1 or 'select' in fcodes_to_save:
                        # Don't save the code if it's just an Operate; that's the default.
                        step['fcodes'] = fcodes_to_save
            action = extra_data_for_function_step(function_name, step_number, 'action')
            if action:
                step['action'] = action
            func['steps'].append(step)
        else:
            print('\tMissing definition for point name {} in function {} step {}'.format(point_name,
                                                                                         function_name,
                                                                                         step_number))

    def load_function(self, csv_input_file, function_name):
        """Construct and return a function from the CSV data in the worksheet named function_name."""
        current_function = {
            'id': function_name,
            'name': function_name,
            'steps': []
        }
        reference = extra_data_for_function(function_name, 'ref')
        if reference is not None:
            current_function['ref'] = reference

        support_point_name = extra_data_for_function(function_name, 'support_point')
        if support_point_name is not None:
            current_function['support_point'] = support_point_name

        for row_num, row in enumerate(csv.DictReader(csv_input_file)):
            try:
                self.create_function_step(current_function, row)
            except ValueError as exc:
                print('\tBad data in row {} of {}: exc={}'.format(row_num, function_name, exc))
        self.function_defs[function_name] = current_function
        return current_function

    def add_points_for_added_function(self, points):
        """Hand-create point definitions for a function."""
        edit_selector_name = points['edit_selector_name'] if 'edit_selector_name' in points else None
        if edit_selector_name:
            edit_selector_start = self.next_index['AO']
            edit_selector_point = self.create_extra_output_point('AO', edit_selector_name)
            edit_selector_point['type'] = 'selector_block'
        else:
            edit_selector_start = None
            edit_selector_point = None
        for point_name in points['other_point_names']:
            self.create_extra_output_point('AO', point_name)
        array_head_point_name = points['array_point_name'] if 'array_point_name' in points else None
        if array_head_point_name:
            array_point = self.create_extra_output_point('AO', array_head_point_name)
            array_point['type'] = 'array'
            array_point['array_points'] = points['array_columns']
            array_point['array_times_repeated'] = MAX_ARRAY_POINTS
            # Make room in the indexes for the array's points
            self.next_index['AO'] += 2 * MAX_ARRAY_POINTS - 1

        if edit_selector_point:
            edit_selector_point['selector_block_start'] = edit_selector_start
            next_index = self.get_next_index('AO')
            edit_selector_point['selector_block_end'] = next_index - 1
            self.next_index['AO'] = next_index      # Don't advance the index unnecessarily

    def create_mesa_points(self):
        """Read input data. Create and return a JSON list of point definitions."""
        self.create_point_csvs_from_excel()
        for data_type_name, default_row_values in DATA_TYPES.items():
            with open(data_type_name + '.csv', 'rU') as input_file:
                for row_num, row in enumerate(csv.DictReader(input_file)):
                    if row_num <= LAST_ROW[data_type_name]:
                        # Skip high-numbered rows that have useless emptiness or unsupported data
                        self.load_point(data_type_name, row_num, row, default_row_values)

        print('Finished with initial point creation')
        # Create additional points for names that appear in function definitions but not in the spreadsheet.
        for data_type_name, additional_point_names in MORE_POINTS.iteritems():
            for additional_point_name in additional_point_names:
                self.create_extra_output_point(data_type_name, additional_point_name)
        self.add_points_for_added_function(CURVE_FUNCTION_POINTS)
        self.add_points_for_added_function(SCHEDULE_FUNCTION_POINTS)

        for data_type_name in DATA_TYPES.keys():
            os.remove('{}.csv'.format(data_type_name))

    def create_mesa_functions(self):
        """Read input data, transforming into a yaml file of function definitions."""
        workbook = load_workbook(FUNCTION_INPUT_FILE, data_only=True)
        for sheet_name in workbook.sheetnames:
            worksheet_name = sheet_name.encode('ascii', 'ignore')
            with open(''.join(['tmp.csv']), 'wb') as csv_output_file:
                wr = csv.writer(csv_output_file, quoting=csv.QUOTE_ALL)
                for idx, row in enumerate(workbook[worksheet_name].iter_rows()):
                    wr.writerow([self.scrubbed_cell_value(cell.value) for cell in row])
            with open('tmp.csv', 'rU') as csv_input_file:
                self.function_defs[worksheet_name] = self.load_function(csv_input_file, worksheet_name)
            os.remove('tmp.csv')

    def sort_points(self):
        """For easier readability, sort the resulting list of definitions by data_type and index."""
        self.point_defs = sorted(self.point_defs, key=lambda k: (k['data_type'], k['index']))

    def validate_points(self):
        """
            Validate the resulting point definitions:
                Point indexes must be unique within each point type
                All names must be unique except within arrays, in which internal point names = head point name

            Throw an exception (thus preventing the output file write) if validation fails.
        """
        # Validation #1: All indexes must be unique within a data type.
        for dt in DATA_TYPES.keys():
            temp_dict = {}
            for pt in self.point_defs:
                if pt['data_type'] == dt:
                    ind = pt['index']
                    if ind in temp_dict:
                        error_msg = 'Found a duplicate index {} for data type {}'
                        raise ValueError(error_msg.format(ind, dt))
                    temp_dict[ind] = pt

        # Validation #2: All names must be unique, except that points in an array share a common name.
        for pt in self.point_defs:
            pt_by_name = self.points_by_name[pt['name']]
            if pt_by_name['index'] != pt['index']:
                error_msg = 'Found duplicate points: {}.{} ({}), {}.{} ({})'
                raise ValueError(error_msg.format(pt['data_type'], pt['index'], pt['name'],
                                                  pt_by_name['data_type'], pt_by_name['index'], pt_by_name['name']))

    def validate_functions(self):
        """
            Validate the results:
                1. No point may be referenced more than once by functions.
                2. All points referenced by functions must be output points.

            Throw an exception (thus preventing the output file write) if validation fails.
        """
        point_function_dict = {}
        for function_name, func in self.function_defs.iteritems():
            for step in func['steps']:
                # 1. Confirm that the point is referenced by at most one function.
                point_name = step['point_name']
                if point_name in point_function_dict:
                    other_function_name = point_function_dict[point_name]['name']
                    raise ValueError('Found a duplicate use of {} in {} and {}'.format(point_name,
                                                                                       function_name,
                                                                                       other_function_name))
                else:
                    point_function_dict[point_name] = func
                # 2. Confirm that the point is an output point.
                pt = self.points_by_name[point_name]
                if pt['group'] != 40 and pt['group'] != 10:
                    raise ValueError('Found an input point {} in {}'.format(point_name, function_name))

    def save_point_defs(self):
        if self.point_defs:
            with open(POINT_OUTPUT_FILE, 'w') as output_file:
                output_file.write(json.dumps(self.point_defs, sort_keys=True, indent=4))
                print('Created point definition file {}'.format(POINT_OUTPUT_FILE))

    def save_function_defs(self):
        if self.function_defs:
            functions_to_write = {'functions': self.function_defs.values()}
            with open(FUNCTION_JSON_OUTPUT_FILE, 'w') as output_file:
                output_file.write(json.dumps(functions_to_write, sort_keys=True, indent=4))
            with open(FUNCTION_YAML_OUTPUT_FILE, 'w') as output_file:
                import yaml
                output_file.write(yaml.dump(functions_to_write, default_flow_style=False))
            print('Created function definition files {} and {}'.format(FUNCTION_JSON_OUTPUT_FILE,
                                                                       FUNCTION_YAML_OUTPUT_FILE))

    def run(self):
        self.create_mesa_points()
        self.create_mesa_functions()
        self.sort_points()
        self.validate_points()
        self.validate_functions()
        self.save_point_defs()
        self.save_function_defs()


if __name__ == '__main__':
    Converter().run()
